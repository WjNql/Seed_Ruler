# -*- coding: utf-8 -*-
import numpy as np
from scipy.linalg import eig
import openpyxl
from scipy.linalg import eigh

def FitEllip(X, Y, N):
    # Open the Excel file
    workbook = openpyxl.load_workbook('D:\\B_GENERAL\\coding transform\\project\\matlabGUI\\data3.xlsx')

    # Select the worksheet to read
    worksheet = workbook['Sheet1']

    # Create an empty array to store data
    X_data = []

    # Iterate through each row in the worksheet
    for row in worksheet.iter_rows(values_only=True):
        X_data.append(row)
    X = np.array(X_data).flatten()

    # Open another Excel file
    workbook = openpyxl.load_workbook('D:\\B_GENERAL\\coding transform\\project\\matlabGUI\\data4.xlsx')

    # Select the worksheet to read
    worksheet = workbook['Sheet1']

    # Create an empty array to store data
    Y_data = []

    # Iterate through each row in the worksheet
    for row in worksheet.iter_rows(values_only=True):
        Y_data.append(row)
    Y = np.array(Y_data).flatten()

    N = 720
    mx = np.mean(X)
    my = np.mean(Y)
    sx = (np.max(X) - np.min(X)) / 2
    sy = (np.max(Y) - np.min(Y)) / 2
    x = (X - mx) / sx
    y = (Y - my) / sy

    D = np.column_stack((x*x, x*y, y*y, x, y, np.ones_like(x)))
    S = np.dot(D.T, D)

    C = np.zeros((6, 6))
    C[0, 2] = -2
    C[1, 1] = 1
    C[2, 0] = -2

    workbook = openpyxl.load_workbook('D:\\B_GENERAL\\coding transform\\project\\matlabGUI\\data5.xlsx')

    # Select the worksheet to read
    worksheet = workbook['Sheet1']

    # Create an empty array to store data
    S_data = []

    # Iterate through each row in the worksheet
    for row in worksheet.iter_rows(values_only=True):
        S_data.append(row)
    S = np.array(S_data)

    geval, gevec = eig(S, C)

    # Sort the eigenvectors
    sort_indices = np.argsort(geval)[::-1]
    geval = geval[sort_indices]
    gevec = gevec[:, sort_indices]
    neg_indices = np.where((geval < 0) & (~np.isinf(geval)))[0]

    # Extract eigenvectors corresponding to negative eigenvalues
    A = gevec[:, neg_indices]

    print("gevec", gevec)
    print("geval", geval)
    print('A', A)
    a = np.zeros(6)

    a[0] = A[0] * sy * sy
    a[1] = A[1] * sx * sy
    a[2] = A[2] * sx * sx
    a[3] = -2 * A[0] * sy * sy * mx - A[1] * sx * sy * my + A[3] * sx * sy * sy
    a[4] = -A[1] * sx * sy * mx - 2 * A[2] * sx * sx * my + A[5] * sx * sx * sy
    a[5] = A[0] * sy * sy * mx * mx + A[1] * sx * sy * mx * my + A[2] * sx * sx * my * my - A[3] * sx * sy * sy * mx - A[4] * sx * sx * sy * my + A[5] * sx * sx * sy * sy
    a = a.reshape(-1, 1)

    theta = np.arctan2(np.real(a[1]), np.real(a[0] - a[2])) / 2
    ct = np.cos(theta)
    st = np.sin(theta)
    ap = a[0] * ct * ct + a[1] * ct * st + a[2] * st * st
    cp = a[0] * st * st - a[1] * ct * st + a[2] * ct * ct
    T = np.array([[a[0], a[1]/2], [a[1]/2, a[2]]])
    T = np.squeeze(T)

    t = -np.linalg.inv(2*T) @ np.reshape(np.array([a[3], a[4]]), (2, 1))

    cx = t[0]
    cy = t[1]
    val = t.T @ T @ t
    scale = 1 / (val - a[5])

    r1 = 1 / np.sqrt(scale * ap)
    r2 = 1 / np.sqrt(scale * cp)

    v = np.array([r1, r2, cx, cy, theta], dtype=object).reshape(-1, 1)

    if r1 < r2:
        v = np.array([r2, r1, cx, cy, theta + np.pi/2]).reshape(-1, 1)
    v = np.squeeze(v)
    print("r1", r1)
    print("r2", r2)
    print("v", v)
    dx = 2 * np.pi / N
    elliptheta = v[4]
    cos_theta = np.cos(elliptheta)
    sin_theta = np.sin(elliptheta)
    Rad = np.array([[cos_theta, sin_theta], [-sin_theta, cos_theta]]).T

    ellipX = np.zeros(N)
    ellipY = np.zeros(N)
    for i in range(N):
        ang = i * dx
        x = v[0] * np.cos(ang)
        y = v[1] * np.sin(ang)
        d11 = np.matmul(Rad, np.array([x, y]))

        ellipX[i] = d11[0] + v[2]
        ellipY[i] = d11[1] + v[3]

    newX = ellipX
    newY = ellipY

    return newX, newY, v
